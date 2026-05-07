#!/usr/bin/env python3
"""Extract student credentials from the Excel sheet into students.json."""
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "ايميلات الطلبة AI.xlsx"
DST = Path(__file__).resolve().parent / "students.json"
SHEET = "قاعة H5"


def main():
    if not XLSX.exists():
        print(f"ERROR: workbook not found: {XLSX}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f"ERROR: sheet '{SHEET}' not found. Available: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)
    ws = wb[SHEET]

    students = []
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        if not isinstance(row[0], int):
            continue
        num, name, email, pw = row[0], row[1], row[2], row[3]
        if name and email and pw:
            students.append({
                "id": num,
                "name": str(name).strip(),
                "email": str(email).strip().lower(),
                "password": str(pw).strip(),
            })

    DST.write_text(
        json.dumps({"students": students}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"Wrote {len(students)} students → {DST}")


if __name__ == "__main__":
    main()
